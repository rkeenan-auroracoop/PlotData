import os
import pprint
# import openpyxl
from openpyxl import load_workbook

docDir = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry'

harvestWriteFile = (
    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(HARVEST).txt', 'w')

readFile = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\2020 Aaron Franson Test Plot.xlsx'

wb = load_workbook(readFile, data_only=True)
harvestSheet = wb['HARVEST FORM']


def topHarvestInfo():
    GROWER_NAME = harvestSheet['C3'].value
    GROWER_CITY = harvestSheet['C4'].value
    COUNTY = harvestSheet['C5'].value
    ACE_LOCATION = harvestSheet['C6'].value
    STATED_PLOT_ON = harvestSheet['C7'].value
    FLAT_LOCATION = harvestSheet['C8'].value
    GPS_LATITUDE = harvestSheet['C9'].value
    FUNGICIDE = harvestSheet['C10'].value
    if FUNGICIDE != None:
        FUNGICIDE = FUNGICIDE
    else:
        FUNGICIDE = "None"
    HARVEST_DATE = harvestSheet['C11'].value
    if HARVEST_DATE != None:
        HARVEST_DATE = HARVEST_DATE
    else:
        HARVEST_DATE = "None"
    CROP = harvestSheet['H3'].value
    PLANTING_DATE = harvestSheet['H4'].value
    SEEDING_RATE = harvestSheet['H5'].value
    PLANTING_DEPTH_IN = harvestSheet['H6'].value
    PLANTER_TYPE = harvestSheet['H7'].value
    ROW_WIDTH = harvestSheet['H8'].value
    GPS_LONGITUDE = harvestSheet['H9'].value
    HERBICIDE = harvestSheet['H10'].value
    if HERBICIDE != None:
        HERBICIDE = HERBICIDE
    else:
        HERBICIDE = "None"
    COMMODITY_PRICE = harvestSheet['H11'].value
    PLOT_TYPE = harvestSheet['L5'].value
    IRRIGATION_TYPE = harvestSheet['L6'].value
    PREVIOUS_CROP = harvestSheet['L7'].value
    TILLAGE_SYSTEM = harvestSheet['L8'].value
    SOIL_TEXTURE = harvestSheet['L9'].value
    INSECTICIDE_RATE = harvestSheet['L10'].value
    if INSECTICIDE_RATE != None:
        INSECTICIDE_RATE = INSECTICIDE_RATE
    else:
        INSECTICIDE_RATE = "None"
    DRYING_COST = harvestSheet['L11'].value
    FORM_TYPE = "HARVEST FORM"

    f1.write(GROWER_NAME + "\t" + GROWER_CITY + "\t" + COUNTY + "\t" + ACE_LOCATION + "\t" + STATED_PLOT_ON + "\t" + FLAT_LOCATION + "\t" + str(GPS_LATITUDE) + "\t" + FUNGICIDE + "\t" + HARVEST_DATE + "\t" + CROP + "\t" + str(PLANTING_DATE) + "\t" + str(SEEDING_RATE) + "\t" +
                           str(PLANTING_DEPTH_IN) + "\t" + PLANTER_TYPE + "\t" + str(ROW_WIDTH) + "\t" + str(GPS_LONGITUDE) + "\t" + HERBICIDE + "\t" + str(COMMODITY_PRICE) + "\t" + PLOT_TYPE + "\t" + IRRIGATION_TYPE + "\t" + PREVIOUS_CROP + "\t" + TILLAGE_SYSTEM + "\t" + SOIL_TEXTURE + "\t" + INSECTICIDE_RATE + "\t" + str(DRYING_COST) + "\t" + FORM_TYPE + "\t")


def bottomHarvestInfo():
    for row in range(17, harvestSheet.max_row + 1):
        BRAND = harvestSheet['A' + str(row)].value
        if BRAND != None:
            BRAND = BRAND
        else:
            BRAND = 'None'
        PRODUCT = harvestSheet['C' + str(row)].value
        if PRODUCT != None:
            PRODUCT = PRODUCT
        else:
            PRODUCT = 'None'
        ROW_LENGTH = harvestSheet['F' + str(row)].value
        if ROW_LENGTH != None:
            ROW_LENGTH = ROW_LENGTH
        else:
            ROW_LENGTH = "None"
        WET_WEIGHT = harvestSheet['G' + str(row)].value
        if WET_WEIGHT != None:
            WET_WEIGHT = WET_WEIGHT
        else:
            WET_WEIGHT = "None"
        HARVEST_MOISTURE_PCT = harvestSheet['H' + str(row)].value
        if HARVEST_MOISTURE_PCT != None:
            HARVEST_MOISTURE_PCT = HARVEST_MOISTURE_PCT
        else:
            HARVEST_MOISTURE_PCT = "None"
        NUM_OF_ROWS = harvestSheet['I' + str(row)].value
        TEST_WEIGHT = harvestSheet['J' + str(row)].value
        if TEST_WEIGHT != None:
            TEST_WEIGHT = TEST_WEIGHT
        else:
            TEST_WEIGHT = "None"
        YIELD_PER_ACRE = harvestSheet['K' + str(row)].value
        if YIELD_PER_ACRE != None:
            YIELD_PER_ACRE = YIELD_PER_ACRE
        else:
            YIELD_PER_ACRE = "None"
        PCT_OF_PLOT_ADVANTAGE = harvestSheet['L' + str(row)].value
        if PCT_OF_PLOT_ADVANTAGE != None:
            PCT_OF_PLOT_ADVANTAGE = PCT_OF_PLOT_ADVANTAGE
        else:
            PCT_OF_PLOT_ADVANTAGE = "None"
        YIELD_PER_ACRE_RANK = harvestSheet['M' + str(row)].value
        if YIELD_PER_ACRE_RANK != None:
            YIELD_PER_ACRE_RANK = YIELD_PER_ACRE_RANK
        else:
            YIELD_PER_ACRE_RANK = "None"
        DOLLARS_PER_ACRE_RANK = harvestSheet['N' + str(row)].value
        if DOLLARS_PER_ACRE_RANK != None:
            DOLLARS_PER_ACRE_RANK = DOLLARS_PER_ACRE_RANK
        else:
            DOLLARS_PER_ACRE_RANK = "None"
        if ROW_LENGTH != 0 and NUM_OF_ROWS != 0:
            topHarvestInfo()
            f1.write(str(BRAND) + "\t" + str(PRODUCT) + "\t" + str(ROW_LENGTH) +
                     "\t" + str(WET_WEIGHT) + "\t" + str(HARVEST_MOISTURE_PCT) + "\t" + str(NUM_OF_ROWS) + "\t" + str(TEST_WEIGHT) + "\t" + str(YIELD_PER_ACRE) + "\t" + str(PCT_OF_PLOT_ADVANTAGE) + "\t" + str(YIELD_PER_ACRE_RANK) + "\t" + str(DOLLARS_PER_ACRE_RANK) + "\n")
        else:
            break

    print("Your Harvest Form data plot file is done!")


with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(HARVEST).txt', 'w') as f1:
    bottomHarvestInfo()
    wb.close()
