import os
import pprint
# import openpyxl
from openpyxl import load_workbook

# This is for the PLANTING FORM

docDir = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry'

plantingWriteFile = (
    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(PLANTING).txt', 'w')

PV_113_V89 = ['113-V89', '113-V89 VT2', '113-V89 VT2P', '113-V89 VT2PRIB', '113-V89VT2', '113-V89VT2PRIB', '113V89 VT2', 'PV 113 V89 VT2', 'PV 113-V89',
              'PV 113-V89 - VT2', 'PV 113-V89 VT2PRIB 32k', 'PV113 V89 VT2PRIB', 'PV113-V89', 'PV113-V89 VT2', 'PV113-V89 VT2RIB- Check', 'PV113-V89-VT2PRIB', 'PV113-V89VT2']

NK_1082_5222A = ['1082 5222', '1082-5222', '1082-5222A', '1082A 5222',
                 'NK1082-5222A', 'NK1082-5222A Brand', 'NK1082-5222AEZ1', 'NK1082 5222A', 'NK1082 5222']

PV_110_H20SS = ['110-H20 SS', '110-H2O STX']

PV_110_H20VT2PRIB = ['110-H20 VT2', '110-H20 VT2P', '110-H20 VT2PRIB', '110-H20VT2', '110-H20VT2PRIB', 'PV110 H20 VT2PRIB',
                     'PV110-H20 VT2', 'PV110-H20 VT2PRIB', 'PV110-H20 VT2PRIB 34k', 'PV110-H20 VT2RIB', 'PV110-H20VT2', 'PV110 H2O VT2PRIB', 'PV110-H20-VT2PRIB', 'PV110H20-VT2']
PV_112_Q70BT2PRIB = ['112-Q70', '112-Q70 VT2',
                     '112-Q70 VT2P', '112-Q70 VT2PRIB', '112-Q70VT2', 'PV 112- Q70', 'PV 112-Q70 - VT2', 'PV 112-Q70 VT2', 'PV112 Q70 VT2PRIB', 'PV112-Q70', 'PV112-Q70 VT2', 'PV112-Q70 VT2PRIB', 'PV112-Q70 VT2PRIB 36k', 'PV112-Q70TT2', 'PV112Q-70']

PV_112_T69SS = ['112-T69 SS', '112-T69 SSTX',
                '112-T69 SSTX RIB', '112-T69 STX', '112-T69SSRIB', 'PV 112-T69 - VT2', 'PV112 T69 SSRIB', 'PV112-T69', 'PV112-T69 SS', 'PV112-T69 SSRIB', 'PV112-T69 SSRIB 36k']
PV_113_B40 = ['113-B40 DGVT2', '113-B40 DGVT2PRIB',
              '113-B40 VT2', '113-B40 VT2P', '113-B40DGVT2', 'PV113 B40 DGVT2PRIB', 'PV113-B40', 'PV113-B40 DGVT2PRIB', 'PV113-B40 VT2', 'PV113-B40DGVT2']
PV_115_D59 = ['115-D59 VT2', '115-D59 VT2P', '115-D59 VT2P', '115-D59VT2PRIB', '115D59',
              '115D59 VT2', 'PV 115-D59 - VT2', 'PV 115-D59 VT2', 'PV 115-D59 VT2PRIB 34k', 'PV 115-D59VT2', 'PV115 D59 VT2PRIB', 'PV115-D59', 'PV115-D59 VT2', 'PV115-D59 VT2PRIB', 'PV115-D59 VT2RIB', 'PV115-D59TV2', 'PV115-D59VT2']
PV_115_M60 = ['115-M60 TRE RIB', '115-M60 TRERIB', '115-M60 Tr', '115-M60TRE RIB', '115-M60TRERIB', '115-R60 Tre', '115M60', '115M60 VT2',
              'PV 115-M60 TRECEPTA 34k', 'PV115 M60 TRERIB', 'PV115-M60', 'PV115-M60 TRE', 'PV115-M60 TRERIB', 'PV115-M60 TreceptraRIB', 'PV115-M60VT2', 'PV115M60 TRERIB']
PV_114_R50_VT2PRIB = ['114-R50 VT2', '114-R50 VT2P',
                      '114-R50VT2PRIB', '114R50 VT2', 'PV 114 R50 VT2', 'PV 114-R50 VT2PRIB 32k', 'PV114 R50 VT2PRIB', 'PV114-R50 VT2', 'PV114-R50 VT2PRIB', 'PV114-R50-VT2RIB', 'PV114-R50VT2', 'PV114R50VT2', 'PV114-R50-VT2PRIB', 'PV114-R50-VT2']
PV_114_R50_SS = ['114-R50 SS', '114-R50 STX',
                 'PV 114 R50SS', 'PV 114-R50 -SSX']
PV_109_P29_VT2PRIB = ['109-P29 VT2', '109-P29VT2PRIB', 'PV 109-P29 - VT2', 'PV109-P29',
                      'PV109-P29 VT2', 'PV109-P29 VT2PRIB', 'PV109-P29 VT2PRIB 34k', 'PV109-P29 Vt2P', 'PV109-P29VT2', ]
PV_109_A90_VT2PRIB = ['109-A90', '109-A90 VT2', '109-A90 VT2P',
                      '109-A90 VT2PRIB', '109-A90VT2', '109-A90VT2PRIB', 'PV 109-A90 VT2', 'PV109-A90', 'PV109-A90VT2', 'PV 109-A90 VT2PRIB']
PV_3519X = ['3519X']
LG_5643_VT2P = ['LG5643 VT2', '5643 VT2',
                '5643VT2', '5643VT2RIB', 'LG5643 VT2', 'LG5643 VT2PRIB', 'LG5643 VT2PRIB']
LG_5643_STX = ['5643 STX', 'LG5643STX', '5643SS-RIB']
LG_5700_VT2P = ['LG5700VT2', '5700 VT2', '5700 VT2RIB',
                '5700VT2', 'LG5700 VT2', 'LG5700VT2', 'LG5700 VT2PRIB', 'LG5700 VT2PRIB']
LG_5700_STX = ['5700 STX', 'LG5700STX']
LG_5525 = ['LG5525']
LG_61C48_VT2P = ['LG61C48VT2P', 'LG61C48 VT2']
LG_62C35_VT2P = ['62C35 NT2', '62C35 VT2', '62C35VT2', '62C35VT2RIB',
                 'LG62C35', 'LG62C35 VT2', 'LG62C35VT2', 'LG62C35VT2P', '62C34']
LG_64C30_TRC = ['LG64C30TRC', '64C30 VIP',
                '64C30 TRE', '64C30 VT2', '64C30TRC']
LG_59C66_VTP2 = ['59C66 VT2', 'LG59C66 VT2']
LG_59C41_STX = ['59C41 STX']
LG_66C32_STX = ['LG66C32STX', '66C32 STX',
                '66C32SSTX', 'LG66C32 STX', 'LG66C32STX']
LG_66C32_VT2P = ['66C32 VT2']
LG_66C28_3110 = ['66C28 - 3110', '66C28-3110', '66C28 3220', '66C28 VIP']
LG_67C45_STX = ['LG 67C45 STX RIB', '67C45 SS-RIB', '67C45STX', '67C45 STX']
LG_60C33_VT2P = ['LG60C33 VT2', '60C33 VT2']
LG_5650_VT2PRIB = ['LG5650 VT2PRIB']
LG_5606_STX = ['5606 STX']
LG_55C95_VT2P = ['55C95 VT2']
LG_58C77_VT2P = ['58C77 3220']
LG_59C72_VT2P = ['59C72 VT2']
LG_68C88_VT2P = ['68C88 VT2']
LG_C2888RX = ['2888RX']
LG_S2989RX = ['2989RX']
LG_S3060RX = ['3060RX']
LG_C3550RX = ['3550RX']
LG_S3600RX = ['3600RX']
LG_59C46 = ['59C46']
LG_64C30_TRC = ['64C30 TRE', '64C30TRC', 'LG64C30TRC',
                'LG6430 TRCIB', '64C30 VIP', '64C30 VT2']
MYCO_2410Q = ['2410 Q', '2410Q']
MYCO_2410AM = ['2410 AML', '2410 AM', '2410-AM',
               '2410AM', 'MY2410AM', 'MYCO 2410 AM']
MYCO_1610Q = ['1610', '1610 SS']
MYCO_1201Q = ['1201 Q', '1201 Qrome', '1201Q', 'MY1201Q']
MYCO_2470AM = ['MY2470AM']
MYCO_2470AML = ['2470 AML', '2470AML',
                'MY2470AML', 'MYCO 2470 AML', 'MYCO2470AML', '2470-AML']
MYCO_2470Q = ['2470Q']
MYCO_MY1404AM = ['MY1404AM', 'MY1404 AM']
MYCO_1404AM = ['1404AM', '1404 AM']
MYCO_1830AML = ['MYCO 1830 AML' '1830', '1830 AML',
                '1830-AML', '1830AML', 'MYCO1830AML']
MYCO_2030Q = ['2030 -Q', '2030Q', 'MYCO 2030Q']
MYCO_1890Q = ['1890 Q', '1890Q']
MYCO_1101Q = ['MY1101 CYFR', 'MY1101Q', '1101 Q', '1101Q']
MYCO_2630AM = ['2630 AM', '2630AM', 'MYCO 2630 AM']
MYCO_2630Q = ['2630 Q', '2630Q', '2360Q']
MYCO_12G35RA = ['MY12G35', '12G35']
MYCO_MY319E = ['MY319E']
MYCO_MY320L4 = ['MY320E', '320-E3']
MYCO_MY340E = ['340E', 'MY340E']
MYCO_MY350L4 = ['MY350E']
MYCO_5N360R2 = ['MY360E']
MYCO_2358AM = ['2358 AM', '2358AM']
MYCO_289E = ['289-E3', '289E', 'MY289E']
MYCO_2580Q = ['2580AM', '2580 AM']
MYCO_2290AM = ['2290 AM']
MYCO_MY300E = ['300-E3', 'MY300E']
NK_1239_5122 = ['NK 1239-5122', 'NK1239', 'NK1239-5122',
                'NK1239-5122 Brand', 'NK1239-5122-EZ1', '1239 5122', '1239-5122', 'NK1239 5122A', 'NK1239 5122']
NK_1460_5222 = ['1460 5222', 'NK 1460-5222', 'NK1460', 'NK1460-5222',
                'NK1460-5222 Brand', 'NK1460-5222-EZ1', 'NKNX11406-5222', '1460-5222', 'NK1460 5222A', '1406-5222', '1460-5222', 'NK1460 5222']
NK_1354_5222 = ['NK1354-5222', 'NK1354-5222-EZ1',
                '1354-5222', 'NK1354-5222', 'NK1354-5222-EZ1']
NK_1573_5222 = ['1573 -5222', '1573-5222',
                'NK 1573-5222', 'NK1573-5222-EZ1', 'NK1573 5222']
NK_1523_3220 = ['NK1523-3220-EZ1', '1523 - 1523',
                '1523 3220', 'NK 1523-3220', 'NK1523', 'NK1523-3220', '1523 5222', 'NK1523 3220A', '1523-3220', 'NK1523 3220']
NK_1694_3111 = ['1694 3111', '1694-3111']
NK_1188_5122 = ['NK 1188-5122', 'NK1188-5122', 'NK1188-5122 Brand',
                '1188 5122', '1188-5122', '1188-5122A', '1188-5222']
NK_0821_5122A = ['NK0821-5122A', '0821 5122']
NK_1354_3220 = ['1354 3220']
NK_S30_E3 = ['NK S30-E3', 'S30 E3', '30-E3']
NK_S31_E3 = ['NK S31-E3S', 'S31 E3', '31-E3']
NK_S35_E3 = ['NK S35-E3', 'S35 E3', '35-E3', 'S35 E3 ']
NK_S28_E3 = ['S28 E3', 'S28-E3', '28-E3']
NK_1026_3330 = ['1026 3330']
NK_S30_M9X = ['30-M9X']
NK_S35_K9X = ['35-K9X']
NK_S37_A4X = ['37-A4X']
NK_S39_G2X = ['39-G2X']
NK_S39_E3 = ['39-E3']
NK_1460_3110 = ['1460-3220']
NK_S33_E3 = ['S33 E3']
P_1108Q = ['1108', '1108 Q', '1108 Qrome', '1108Q', 'P1108Q', 'P1108Q']
P_1082AM = ['1082 AM', '1082AM', 'P1082AM']
P_1082Q = ['1082 Q', '1082Q']
P_1089AM = ['1089 AMXT', 'P1089AM']
P_1093Q = ['1093 Q', 'P1093Q']
P_1185AM = ['P1185', 'P1185 AM', 'P1185AM', 'P1185AM - CHECK', '1185 AM']
P_1185Q = ['P1185Q', '1185Q']
P_1244Q = ['1244 Qrome', '1244', '1244Q']
P_1353AM = ['1359 AM', '1359AM', 'P1353AM']
P_1353Q = ['P1353Q', '1353Q', '1353 Q']
P_1359AM = ['1359 AM', 'P1359', 'P1359AM']
P_1366Q = ['1366 Q', '1366 Qrome', '1366Q', 'P1366Q']
P_1366AM = ['P1366 - AM']
P_1366AML = ['1366AML', 'P1366AML']
P_1370Q = ['P1370Q']
P_1415Q = ['P1415Q']
P_1548AM = ['P1548AM']
P_1563AM = ['1563AM', 'P1563AM']
P_1563AML = ['P1563AML']
P_1572AM = ['P1572AM']
P_1716Q = ['1716Q', 'P1716', 'P1716Q']
P_1828Q = ['1828Q']
P_1828AM = ['P1828AM']
P_1847AML = ['P1847AML']
P_2042AML = ['2042AML', 'P2042AML']
P_2042AM = ['P2042AM']
P_32T26E = ['P32T26 E', 'P32T26E']
P_0622Q = ['P0622Q', 'PO622Q']
P_0339AM = ['PO339AM', 'P0339AM']
P_0343AML = ['P0343AML', 'PO343AML']
P_0446Q = ['P0446Q', 'PO446Q']
P_0595AM = ['P0595AM', 'PO595AM']
AC_3219_E3 = ['A3219E3']
AC_3619_E3 = ['A3619E3', '3619-E3']
AC_3719_E3 = ['A3719E3']
AC_2919_E3 = ['2919-E3', 'A2919E3']
HOEG_7404Q = ['7404Q']
HOEG_7436Q = ['7436Q']
HOEG_7692Q = ['7692Q']
HOEG_7990Q = ['7990Q']
HOEG_8028AM = ['8028 AM']
HOEG_8073Q = ['8073Q']
HOEG_8085Q = ['8085Q']
HOEG_8106Q = ['8106Q']
HOEG_8188Q = ['8188Q']
HOEG_8233AM = ['8233 AM']
HOEG_8235Q = ['8235Q']
HOEG_8268Q = ['8268Q']
HOEG_8364AMXT = ['8364AMXT']
HOEG_8417Q = ['8417Q']
HOEG_8491Q = ['8491Q']
HOEG_8511AML = ['8511 AML']
HOEG_8512Q = ['8512Q']
HOEG_8519Q = ['8519Q']
HOEG_8531Q = ['8531Q']
HOEG_8636AM = ['8636 AM', '8636AM']
HOEG_8637Q = ['8637Q']
HOEG_8749AM = ['8749 AM']
DKC_51_91SS = ['5191', '5191 SS']
DKC_53_27SS = ['5327', '5327 SS']
DKC_54_64SS = ['5464', '5464 SS']
DKC_55_65SS = ['56-65 SS', '5565', '55-65 SS', 'DKC 55-65 SS', 'DKC 55-65 SS']
DKC_59_81SS = ['5981', '59-81 SS', 'DKC59-81RIB', '5981 SS']
DKC_59_82VT2P = ['59-82']
DKC_60_88VT2PRIB = ['DKC60-88RIB']
DKC_61_40SS = ['6140', '61-40', '61-40 SS',
               'DKC61-40', 'DKC61-40RIB', '6140 SS']
DKC_63_60SS = ['DKC63-60']
DKC_63_91VT2PRIB = ['63-91']
DKC_63_90SS = ['6930', '63-90', 'DKC63-90',
               'DKC63-90RIB', '6390 SS', 'DKC 63-90', 'DKC 63-90RIB', '6390-SS']
DKC_64_34SS = ['DKC64-34']
DKC_64_35VT2P = ['64-35']
DKC_66_17SS = ['66-17', '66-17SS', 'DKC66-17']
DKC_66_74SS = ['66-74SS', 'DKC66-74', '66-74 SS']
DKC_70_26SSRIB = ['DKC70-26RIB', 'DKC 70-26', 'DKC 70-26RIB']
DKC_70_27VT2P = ['70-27']
DKC_70_27VT2PRIB = ['DK70-27 RIB']
F_90D623 = ['09D623']
F_09G219 = ['09G219']
F_10D308 = ['10D308']
F_11A637 = ['11A637']
F_12A558 = ['12A558']
F_13A843 = ['13A843']
F_13G519 = ['13G519']
F_14D796 = ['14D796']
F_14G658 = ['14G658']
F_15A410 = ['15A410']
F_15A657 = ['15A657']
F_14A648 = ['14A648']
F_16A826 = ['16A826']
F_09A219 = ['09A219']
F_09A249 = ['09A249']
F_11A637 = ['11A637']
F_17A819 = ['17A819']
CHAN_209_15BT2PRIB = ['209-15VT2']
CHAN_213_19STXRIB = ['213-19 STX', '213-19STX']
CHAN_213_19VT2 = ['213-19VT2']
CHAN_213_93STX = ['213-93', '213-93STX', '213-95STXRIB']
CHAN_214_22 = ['214-22']
CHAN_214_78 = ['214-78', '214-78DGVT2', '214-78 VT2', '214-78VT2']
CHAN_215_60TRERIB = ['214-60 TRI', '215-60 TRI']
CHAN_215_75VT2PRIB = ['215-75', '215-75VT2', '215-75VT2P', '215-75VT2RIB']
CHAN_216_36_CONV = ['216-36']
CHAN_217_76VT2PRIB = ['217-76 VT2']
CHAN_217_76STX = ['217-76STX']
AGRI_644_32_TRIRIB = ['44-32 TRI', '644-32TRICRIB']
AGRI_645_16_STX = ['645-16 STX']
AGRI_6652_VT2PRIB = ['6652 VT2', '6652VT2RIB']
AGRI_639_70STXRIB = ['A639-70STSRIB']
AGRI_6499_STXRIB = ['A6499STXRIB']
AGRI_647_90_STXRIB = ['647-90 STX']
AGRI_6652_STX = ['6652 STX']
AV_3917YHB = ['3917', '3917 YHB']
AV_4509AML = ['4509', '4509 AML']
BC_B13_R70_3000GT = ['BCB13-R70-3000GT']
BC_BC15_H64VT2 = ['BCB15-H64VT2']
CHAM_65A17_SSRIB = ['65A17SSRIB']
CP_5789VT2P = ['CP5789VTP2']
CP_5370SS = ['CP5370VTP2']
DG_54SS74 = ['D54SS74']
DG_54VC14 = ['D54VC14']
DG_57VC17 = ['D57VC17']
DG_S35EN99 = ['S35EN99']
DG_S36ES70 = ['S36ES70']
GH_G10l15_5222 = ['G10L16 - 5222A', 'G10L16-5222A']
GH_G12S75_5122 = ['G12575-5122.0']
GH_G13N18_3111 = ['G13N18-3111']
GH_G13T41_5122 = ['G13T41-3120EZ!']
GH_G15L32_5222 = ['G15L32-5222-EZ1']
HE_5832 = [5832]
HE_5922 = [5944]
HE_6024 = [6024, 'H6024SSRIB']
HE_6332 = [6332]
HE_6524 = ['H6524SSRIB']
HE_6532 = [6532]
IC_6038_3330A = ['IC60 38-3330A', 'IC6038-3330 Brand']
IC_6312_3220 = ['IC6312-3220']
IC_6580_3330 = ['IC6580-3330']
IC_6829-3111 = ['IC6829-3111']
JS_9513_SS = [9616]
LE_9016_GENSSRIB = [9016, '9016 IONRx']
NCP_04_99_VT2RIB = ['04-99VT2RIB']
NCP_07_27_VT2RIB = ['07-27VT2RIB']
NCP_11_15_VT2RIB = ['11-15 VT2RIB', '11-15:', 'NC+ 11-15']
NCP_14_64_VT2RIB = ['14-64 VT2RIB']
NCP_15_65_VT2RIB = ['15-65 VT2RIB', 'NC+15-65', 'NC+15-65']
NCP_12_48_DGVT2PRO = ['NC+ 14-48']
NCP_14_88_VT2PRIB = ['NC+14-88']
RSC_6148_3010A = ['RC6148-3010A']


def topPlantingInfo():
    GROWER_NAME = plantingSheet['C3'].value
    GROWER_CITY = plantingSheet['C4'].value
    COUNTY = plantingSheet['C5'].value
    ACE_LOCATION = plantingSheet['C6'].value
    if ACE_LOCATION != None:
        ACE_LOCATION = ACE_LOCATION
    else:
        ACE_LOCATION = "None"
    STARTED_PLOT_ON = plantingSheet['C7'].value
    if STARTED_PLOT_ON != None:
        STARTED_PLOT_ON = STARTED_PLOT_ON
    else:
        STARTED_PLOT_ON = 'None'
    FLAG_LOCATION = plantingSheet['C8'].value
    if FLAG_LOCATION != None:
        FLAG_LOCATION = FLAG_LOCATION
    else:
        FLAG_LOCATION = "None"
    GPS_LATITUDE = plantingSheet['C9'].value
    FUNGICIDE = plantingSheet['C10'].value
    if FUNGICIDE != None:
        FUNGICIDE = FUNGICIDE
    else:
        FUNGICIDE = "None"
    CROP = plantingSheet['H3'].value
    PLANTING_DATE = plantingSheet['H4'].value
    SEEDING_RATE = plantingSheet['H5'].value
    PLANTING_DEPTH_IN = plantingSheet['H6'].value
    if PLANTING_DEPTH_IN != None:
        PLANTING_DEPTH_IN = PLANTING_DEPTH_IN
    else:
        PLANTING_DEPTH_IN = "None"
    PLANTER_TYPE = plantingSheet['H7'].value
    if PLANTER_TYPE != None:
        PLANTER_TYPE = PLANTER_TYPE
    else:
        PLANTER_TYPE = "None"
    ROW_WIDTH = plantingSheet['H8'].value
    if ROW_WIDTH != None:
        ROW_WIDTH = ROW_WIDTH
    else:
        SOIL_TEXTURE = "None"
    GPS_LONGITUDE = plantingSheet['H9'].value
    HERBICIDE = plantingSheet['H10'].value
    if HERBICIDE != None:
        HERBICIDE = HERBICIDE
    else:
        HERBICIDE = "None"
    PLOT_TYPE = plantingSheet['L5'].value
    if PLOT_TYPE != None:
        PLOT_TYPE = PLOT_TYPE
    else:
        PLOT_TYPE = "None"
    IRRIGATION_TYPE = plantingSheet['L6'].value
    if IRRIGATION_TYPE != None:
        IRRIGATION_TYPE = IRRIGATION_TYPE
    else:
        IRRIGATION_TYPE = "None"
    PREVIOUS_CROP = plantingSheet['L7'].value
    if PREVIOUS_CROP != None:
        PREVIOUS_CROP = PREVIOUS_CROP
    else:
        PREVIOUS_CROP = "None"
    TILLAGE_SYSTEM = plantingSheet['L8'].value
    if TILLAGE_SYSTEM != None:
        TILLAGE_SYSTEM = TILLAGE_SYSTEM
    else:
        TILLAGE_SYSTEM = "None"
    SOIL_TEXTURE = plantingSheet['L9'].value
    if SOIL_TEXTURE != None:
        SOIL_TEXTURE = SOIL_TEXTURE
    else:
        SOIL_TEXTURE = "None"
    INSECTICIDE_RATE = plantingSheet['L10'].value
    if INSECTICIDE_RATE != None:
        INSECTICIDE_RATE = INSECTICIDE_RATE
    else:
        INSECTICIDE_RATE = "None"
    FORM_TYPE = "PLANTING FORM"

    f1.write(GROWER_NAME.title() + "\t" + GROWER_CITY.title() + "\t" + COUNTY.title() + "\t" + ACE_LOCATION.title() + "\t" + STARTED_PLOT_ON.title() + "\t" + FLAG_LOCATION.title() + "\t" + str(GPS_LATITUDE) + "\t" + FUNGICIDE.title() + "\t" + CROP.title() + "\t" + str(PLANTING_DATE) + "\t" + str(SEEDING_RATE) + "\t" +
             str(PLANTING_DEPTH_IN) + "\t" + PLANTER_TYPE.title() + "\t" + str(ROW_WIDTH) + "\t" + str(GPS_LONGITUDE) + "\t" + HERBICIDE.title() + "\t" + PLOT_TYPE.title() + "\t" + IRRIGATION_TYPE.title() + "\t" + PREVIOUS_CROP.title() + "\t" + TILLAGE_SYSTEM.title() + "\t" + SOIL_TEXTURE.title() + "\t" + INSECTICIDE_RATE + "\t" + FORM_TYPE.title() + "\t")


def bottomPlantingInfo():

    for row in range(17, plantingSheet.max_row + 1):
        ENTRY = plantingSheet['A' + str(row)].value
        COMPANY = plantingSheet['C' + str(row)].value
        if COMPANY == 'LG':
            COMPANY = 'LG Seeds'
        else:
            COMPANY = COMPANY

        BASEITEMGUID = ''
        HYBRID_VARIETY = plantingSheet['F' + str(row)].value
        for product in PV_113_V89:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 113-V89 VT2PRIB'
                BASEITEMGUID = '8F552B16-DB63-48B0-8C99-DB6E968B1E22'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1082_5222A:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1082-5222A'
                BASEITEMGUID = '047B6471-E671-434C-AF2D-1DE5657F2AEA'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_110_H20SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 110-H20 SS'
                BASEITEMGUID = '22E1DB5A-F533-4CE2-BF66-153706205D5F'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_110_H20VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 110-H20 VT2PRIB'
                BASEITEMGUID = '00B96AF0-0394-436A-A385-978A1C21BD89'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_112_Q70BT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 112-Q70 VT2PRIB'
                BASEITEMGUID = '608CB2B2-9219-414A-8486-EFC9C2542C86'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_112_T69SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 112-T69 SS RIB'
                BASEITEMGUID = '993A89E8-E3F7-47AE-9908-FE7CAAF9D6F7'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_113_B40:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 113-B40 DGVT2PRIB'
                BASEITEMGUID = 'B376985F-E07A-474F-8170-70B1A999C8B2'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_115_D59:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 115-D59 VT2PRIB'
                BASEITEMGUID = 'D27B9A20-13F0-4CFC-84EF-19ECE9B64924'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_115_M60:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 115-M60 TRECEPTA'
                BASEITEMGUID = 'F6B5F7F0-10CD-49D1-9192-57E84EE753EC'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_114_R50_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 114-R50 VT2PRIB'
                BASEITEMGUID = '9B16561A-D598-4F6F-BADA-AC91C027572B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_114_R50_SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 114-R50 SS'
                BASEITEMGUID = '4B9042A1-9226-4BA1-A8E7-9AC8B16600C5'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_109_P29_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 109-P29 VT2PRIB'
                BASEITEMGUID = '94C559B2-C262-43B0-BDA3-E98F7EAC76CE'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_109_A90_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 109-A90 VT2PRIB'
                BASEITEMGUID = '3788CA80-FFEE-4898-9052-DFED4D0A1F3B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in PV_3519X:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'PV 3519X'
                BASEITEMGUID = 'C88D2F86-62D5-476F-AE33-0D6603127C7B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5643_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5643 VT2P RIB'
                BASEITEMGUID = 'BD58062A-D5BC-4C71-9EB9-BABF52E8B384'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5643_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5643 STX RIB'
                BASEITEMGUID = 'D524D3CE-82B8-4492-80AE-07663301546D'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5700_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5700 VT2P RIB'
                BASEITEMGUID = '7B649F45-2512-474A-89FB-5D123DF4E1B6'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5700_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5700 STX RIB'
                BASEITEMGUID = '4F6A0CAE-A6E7-4847-8361-D880BF29DE1C'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5525:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5525'
                BASEITEMGUID = '78AE89F4-83D0-49FD-B890-2A0AD3849B5E'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_61C48_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 61C48 VT2P RIB'
                BASEITEMGUID = 'BE770144-C9C3-4EB7-85FD-3C3CC4A92532'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_62C35_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 62C35 VT2P RIB'
                BASEITEMGUID = '597DE6C5-8C4A-4A3A-B89B-27BB5BAFAB96'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_64C30_TRC:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 64C30 TRC RIB'
                BASEITEMGUID = 'E4172F9A-F63B-448A-929D-5ADF65115C40'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_59C66_VTP2:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 59C66 VT2P RIB'
                BASEITEMGUID = '6531D4F6-9965-4A44-88E7-0E506A7DC5FC'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_59C41_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 59C41 STX RIB'
                BASEITEMGUID = 'C76FE046-C980-48D3-BBD0-F8E84FE51D68'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_66C32_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 66C32 STX RIB'
                BASEITEMGUID = 'F58A604D-0215-40B0-8ABD-6A2BA3D12F05'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_66C32_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 66C32 VT2P RIB'
                BASEITEMGUID = 'D6CF601F-89C8-4663-A155-AA52903ED565'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_66C28_3110:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 66C28 3110'
                BASEITEMGUID = 'BBF34E36-114B-4C4C-8B5F-CB661D44A704'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_67C45_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 67C45 STX RIB'
                BASEITEMGUID = 'DA404D4C-0325-46D9-8E6E-32262E50CB17'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_60C33_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 60C33 VT2P RIB'
                BASEITEMGUID = '59767244-D7E8-4887-B141-B07F5A87269A'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5650_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5650 VT2P RIB'
                BASEITEMGUID = '5873E176-BC43-4B29-A144-7CC73F779996'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_5606_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 5606 STX RIB'
                BASEITEMGUID = '9A73674F-E03D-4D07-970C-AC70FE0F4EB7'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_55C95_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 55C95 VT2P'
                BASEITEMGUID = '89226E37-7A3B-46B1-B817-29D110347AA6'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_58C77_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 58C77 VT2P RIB'
                BASEITEMGUID = 'A27CC89D-7177-41DC-AB41-FB77AE7E77F9'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_59C72_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 59C72 VT2P RIB'
                BASEITEMGUID = '951D5426-254E-4582-B652-0D43F5E8D93D'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_68C88_VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 68C88 VT2P RIB'
                BASEITEMGUID = '29E7DE18-AC94-4ACF-89F5-A79F35974A77'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_C2888RX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG C2888RX'
                BASEITEMGUID = 'DBD6A924-13B6-4EEA-B30F-62D7CFD76924'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_S2989RX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG S2989RX'
                BASEITEMGUID = '65C6770F-EA57-403E-BDF6-8FDA37746E6D'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_S3060RX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG S3060RX'
                BASEITEMGUID = '5A2A26F5-0BD1-4778-91F9-4FF646B0AD6E'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_C3550RX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG C3550RX'
                BASEITEMGUID = '19A4C5B6-E434-4DD3-99BF-7DA444457779'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_S3600RX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG S3600RX'
                BASEITEMGUID = '0CC3D44A-E050-4336-B0D7-E83328067EA5'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_59C46:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 59C46 Conv'
                BASEITEMGUID = '4DE8DBBC-FD1D-46AE-A27F-97CDC7FA1084'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LG_64C30_TRC:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LG 64C30 TRC RIB'
                BASEITEMGUID = 'E4172F9A-F63B-448A-929D-5ADF65115C40'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2410Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2410Q'
                BASEITEMGUID = '15DCB22C-4426-4611-9118-270AD62B6799'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2410AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2410AM'
                BASEITEMGUID = '2E351241-17AC-4B98-BD1F-FA105E10B803'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1610Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1610Q'
                BASEITEMGUID = '878B3D71-D1D4-43C6-A1AA-F2CA6B22C7FF'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1201Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1201Q'
                BASEITEMGUID = '804D4E61-1A96-47CB-9559-F298FC4D7E2C'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2470AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2470AM'
                BASEITEMGUID = '4EB2A823-A071-4A7E-97AF-76F269730B9B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2470AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2470AML'
                BASEITEMGUID = 'ADDFC5AB-B968-4B57-8240-D8A4AFFEA2EA'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2470Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2470Q'
                BASEITEMGUID = '39671E4E-80D4-4CB6-B9B9-D16A0DBE542F'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY1404AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY1404AM'
                BASEITEMGUID = 'B5B2A4AA-1B53-447F-A93F-5E79B1275DA6'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1404AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1404AM'
                BASEITEMGUID = '378F6F09-0882-41C6-9AC7-78B6D662CB89'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1830AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1830AML'
                BASEITEMGUID = '5EB16F04-6313-4EBC-8DD0-1E715E48D7CB'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2030Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2030Q'
                BASEITEMGUID = 'CA3E6FF7-63BD-47F2-A01C-7F8A17963116'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1890Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1890Q'
                BASEITEMGUID = 'F92DEE0D-766E-43E6-857F-FC54A36A672B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_1101Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 1101Q'
                BASEITEMGUID = '15B2DBCF-00E2-436D-B7CA-877105A94CBA'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2630AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2630AM'
                BASEITEMGUID = 'E9DE374A-4223-4621-B9C5-9400ED913BC0'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2630Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2630Q'
                BASEITEMGUID = '275AF472-7DE7-4A19-97ED-D96B76922F12'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_12G35RA:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY12G35RA'
                BASEITEMGUID = 'DC3F238E-BC2D-41BD-9217-50352D9FE394'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY319E:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY319E'
                BASEITEMGUID = '9CCFFC8B-565B-4FA2-9E17-91B4637C4F23'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY320L4:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY320L4'
                BASEITEMGUID = 'D6B213A6-D3E6-436B-A990-0778B4D00DA7'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY340E:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY340E'
                BASEITEMGUID = '774F5B7F-3D52-4E78-9D57-25F6508C36F2'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY350L4:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY350L4'
                BASEITEMGUID = 'C0AAAF8D-E35A-4CF3-A447-2A4627DD6943'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_5N360R2:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 5N360R2'
                BASEITEMGUID = '6159F0DE-4874-47EE-9144-B24542563193'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2358AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2358AM'
                BASEITEMGUID = 'B0D4C565-FF36-476C-A715-3E02B53732EA'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_289E:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY289E'
                BASEITEMGUID = '283CBA00-C109-4903-8C1C-8E0E02A9EEEE'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2580Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = '2580AM'
                BASEITEMGUID = '67489DDD-1B6A-4A16-B79F-9C910F1E3D7D'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_2290AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO 2290AM'
                BASEITEMGUID = 'AF7FE0F9-B0B9-4D09-9A1C-8406AB4680E2'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in MYCO_MY300E:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'MYCO MY300E'
                BASEITEMGUID = 'B5FA8E9B-826B-4D26-94A8-2C21DFCCF4E1'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1239_5122:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1239-5122'
                BASEITEMGUID = '03A74077-2633-42BE-A9C4-3CA0D78596BC'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1460_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1460-5222'
                BASEITEMGUID = '9C7009A2-79CD-4714-B911-B5C33162FB46'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1354_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1354-5222'
                BASEITEMGUID = '04A1AA99-E0D9-487C-8715-773ACE0648CE'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1573_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1573-5222'
                BASEITEMGUID = '8FC0EE4D-93F0-4872-8C75-DCC046A06814'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1523_3220:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1523-3220'
                BASEITEMGUID = '73DB5840-C88D-4224-9D42-F794BC2EB2D2'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1694_3111:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1694-3111'
                BASEITEMGUID = '4083695E-4933-4094-8A8C-E6C8C55A4566'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1188_5122:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1188-5122'
                BASEITEMGUID = 'C6F0C46C-4790-4CDD-AD7A-2FC63227B5D2'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_0821_5122A:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 0821-5122A'
                BASEITEMGUID = 'B1CBC295-C776-40B1-82A0-9D51DD90249B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1354_3220:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1354-3220'
                BASEITEMGUID = 'E8894FBD-839E-48E5-8767-31F60492F001'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1354_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1354-5222'
                BASEITEMGUID = '04A1AA99-E0D9-487C-8715-773ACE0648CE'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S30_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S30-E3'
                BASEITEMGUID = 'D3B77EF8-E919-4683-8EBE-19DA3BD470A0'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S31_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S31-E3'
                BASEITEMGUID = '21F488C0-2A38-4441-BCFB-26CBEBA3BF4D'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S35_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S35-E3'
                BASEITEMGUID = '9807A85B-1173-43DF-BC5D-E858B983F6AB'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S28_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S28-E3'
                BASEITEMGUID = 'B9159B82-7555-4827-AE2C-968BEEFAA74C'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1026_3330:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1026-3330'
                BASEITEMGUID = 'A52D0A6B-06BA-4126-B7B1-1826FA16DEF0'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S30_M9X:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S30-M9X'
                BASEITEMGUID = 'AA5423AF-34AB-43F6-A9F9-9BD6605683C5'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S35_K9X:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S35-K9X'
                BASEITEMGUID = 'D92C4427-1A19-46F2-8300-D7CEAD948182'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S37_A4X:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S37-A4X'
                BASEITEMGUID = 'F900A269-5246-4C61-AF04-B9AE58A422F6'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S39_G2X:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S39-G2X'
                BASEITEMGUID = 'E9D5C4DF-56F9-4BEE-B00E-EB69C9A08D3B'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S30_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S30-E3'
                BASEITEMGUID = 'D3B77EF8-E919-4683-8EBE-19DA3BD470A0'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S39_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S39-E3'
                BASEITEMGUID = '63209980-20D4-4335-846D-91F32EAC7D48'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_1460_3110:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK 1460-3110'
                BASEITEMGUID = 'CC53ED31-5170-44EB-B514-853E2AE9127F'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NK_S33_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NK S33-E3'
                BASEITEMGUID = 'B0A2CE6E-4441-4F25-9EE6-79986BFBBD06'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1108Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1108Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1082AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1082AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1082Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1082Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1089AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1089AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1093Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1093Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1185AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1185AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1185Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1185Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1244Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1244Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1353Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1353Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1353AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1353AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1359AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1359AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1366Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1366Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1366AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1366AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1366AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1366AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1563AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1563AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1563AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1563AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1716Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1716Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1828Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1828Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1828AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1828AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_2042AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P2042AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1370Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1370Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1415Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1415Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1548AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1548AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1572AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1572AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_1847AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P1847AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_2042AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P2042AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_0339AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P0339AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_32T26E:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P32T26E'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_0622Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P0622Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_0343AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P0343AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_0446Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P0446Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in P_0595AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'P0595AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AC_3219_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A3219 E-3'
                BASEITEMGUID = 'D9D58035-7E99-4132-8150-1496662D8B1E'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AC_3619_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A3619 E-3'
                BASEITEMGUID = 'AC05F294-08D4-44FA-9132-DCC6FECA62CC'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AC_3719_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A3719S E-3'
                BASEITEMGUID = 'D3A32B45-531E-4F31-A91B-BFB97C29D6D0'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AC_2919_E3:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A2919 E-3'
                BASEITEMGUID = '94844744-24C1-4465-ADB7-8F017ACCD9DD'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_7404Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 7404Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_7436Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 7436Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_7692Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 7692Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_7990Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 7990Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8028AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8028AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8073Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8073Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8085Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8085Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8106Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8106Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8188Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8188Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8233AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8233AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8235Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8235Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8268Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8268Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8364AMXT:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8364AMXT'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8417Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8417Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8491Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8419Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8511AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8511AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8512Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8512Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8519Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8519Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8531Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8531Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8636AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8636AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8637Q:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8637Q'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HOEG_8749AM:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'Hoeg 8749AM'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_51_91SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 51-91 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_53_27SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 53-27 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_54_64SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 54-64 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_59_81SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 59-81 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_61_40SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 61-40 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_59_82VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 59-82 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_63_91VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKB 63-91 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_64_35VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 64-35 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_66_17SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 66-17 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_70_27VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 70-27 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_70_27VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 70-27 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_60_88VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 60-88 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_63_60SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 63-60 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_64_34SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 64-34 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_66_74SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 66-74 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_70_26SSRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 70-26 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_63_90SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 63-90 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DKC_55_65SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'DKC 55-65 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_90D623:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 09D623 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_09G219:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 09G219 DGVT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_10D308:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 10D308 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_11A637:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 11A637 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_12A558:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 12A558 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_13A843:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 13A843 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_13G519:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 13G519 DGVT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_14D796:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 14D769 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_14G658:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 14G658'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_15A410:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 15A410 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_15A657:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 15A657 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_16A826:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 16A826 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_09A219:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 09A219 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_09A249:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 09A249 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_11A637:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 11A637 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_17A819:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 17A819 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in F_14A648:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'F 14A648 CONV'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_209_15BT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 209-15VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_213_19STXRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 213_19STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_213_19VT2:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 213_19VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_213_93STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 213-93STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_214_78:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 214-78DGVT2'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_215_60TRERIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 215-60TRERIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_215_75VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 215-75VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_216_36_CONV:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 216-36CONV'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_217_76VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 217-76VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_217_76STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 217_76STX'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAN_214_22:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CHAN 214_22'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_644_32_TRIRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A644-32 TRICRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_645_16_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A645-16 STX'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_6652_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A6652 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_639_70STXRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A639-70 STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_6499_STXRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A6499 STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_647_90_STXRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A647-90 STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AGRI_6652_STX:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'A6652 STXRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AV_3917YHB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'AV3917YHB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in AV_4509AML:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'AV4509AML'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in BC_B13_R70_3000GT:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'B13-R70-3000GT'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in BC_BC15_H64VT2:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'B15-H64-VT2'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CHAM_65A17_SSRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = '65A17 SSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CP_5789VT2P:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CP5789VTP2'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in CP_5370SS:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'CP5370VTP2'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DG_54SS74:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'D54SS74'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DG_54VC14:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'D54VC14'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DG_57VC17:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'D57VC17'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in GH_G10l15_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'GH10L16-5222A'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in GH_G12S75_5122:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'GH12S75-5222A'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in GH_G13N18_3111:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'G13N18-3111'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in GH_G13T41_5122:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'G13T41-5122'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in GH_G15L32_5222:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'G15L32-5222'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in DG_S35EN99:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'S35EN99'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_5832:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H5832 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_5922:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H5922 BT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_6024:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H6024 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_6332:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H6332 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_6532:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H6532 VT2P'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in HE_6524:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'H6524 SS'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in IC_6038_3330A:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'IC6038-3330A'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in IC_6312_3220:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'IC6312-3220'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in IC_6580_3330:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'IC6580-3330'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in IC_6829_3111:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'IC6829-3111'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in LE_9016_GENSSRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'LR 9016 GENSSRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_04_99_VT2RIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC 04-99 VT2RIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_07_27_VT2RIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC 07-27 VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_11_15_VT2RIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC 11-15 VT2RIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_14_64_VT2RIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC 14-64 VT2RIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_12_48_DGVT2PRO:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC+ 12-48'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        for product in NCP_14_88_VT2PRIB:
            if HYBRID_VARIETY == product:
                HYBRID_VARIETY = 'NC 14-88VT2PRIB'
                BASEITEMGUID = '14'
            else:
                HYBRID_VARIETY = HYBRID_VARIETY
        SEED_TREATMENTS = plantingSheet['J' + str(row)].value
        if SEED_TREATMENTS != None:
            SEED_TREATMENTS = SEED_TREATMENTS
        else:
            SEED_TREATMENTS = "None"
        NUM_OF_ROWS = plantingSheet['M' + str(row)].value

        if COMPANY != None and HYBRID_VARIETY != None:
            topPlantingInfo()
            f1.write(str(ENTRY) + "\t" + COMPANY.title() + "\t" + str(HYBRID_VARIETY) +
                     "\t" + str(BASEITEMGUID) +
                     "\t" + str(SEED_TREATMENTS) + "\t" + str(NUM_OF_ROWS) + "\n")
        else:
            continue


for folders, sub_folders, file in os.walk(docDir):
    for name in file:
        if name.endswith(".xlsx"):
            filename = os.path.join(folders, name)
            print(filename)
            wb = load_workbook(filename)
            plantingSheet = wb['PLANTING FORM']
            with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(PLANTING).txt', 'a') as f1:
                bottomPlantingInfo()
                wb.close()
        else:
            continue

print("Your Planting Form data plot file is done!")

# with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(PLANTING).txt', 'a+') as f1:
#    for line in f1:
#        if '113-V89' in line:
#            line.replace('113-V89', 'PV 113-V89 VT2PRIB')
#        else:
#            continue

# print("Product names are now updated.")
# wb.close()


# bottomPlantingInfo()
# plantingWriteFile.close()
# wb.close()

# This is for the NOTES FORM

# wb = openpyxl.load_workbook(
#    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\2020 Aaron Franson Test Plot.xlsx')
# notesSheet = wb['NOTES FORM']
# notesWriteFile = open(
#    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(NOTES).txt', 'w')


def topNotesInfo():
    GROWER_NAME = notesSheet['C3'].value
    GROWER_CITY = notesSheet['C4'].value
    COUNTY = notesSheet['C5'].value
    ACE_LOCATION = notesSheet['C6'].value
    STATED_PLOT_ON = notesSheet['C7'].value
    FLAT_LOCATION = notesSheet['C8'].value
    GPS_LATITUDE = notesSheet['C9'].value
    FUNGICIDE = notesSheet['C10'].value
    if FUNGICIDE != None:
        FUNGICIDE = FUNGICIDE
    else:
        FUNGICIDE = "None"
    CROP = notesSheet['H3'].value
    PLANTING_DATE = notesSheet['H4'].value
    SEEDING_RATE = notesSheet['H5'].value
    PLANTING_DEPTH_IN = notesSheet['H6'].value
    PLANTER_TYPE = notesSheet['H7'].value
    ROW_WIDTH = notesSheet['H8'].value
    GPS_LONGITUDE = notesSheet['H9'].value
    HERBICIDE = notesSheet['H10'].value
    if HERBICIDE != None:
        HERBICIDE = HERBICIDE
    else:
        HERBICIDE = "None"
    PLOT_TYPE = notesSheet['L5'].value
    IRRIGATION_TYPE = notesSheet['L6'].value
    PREVIOUS_CROP = notesSheet['L7'].value
    TILLAGE_SYSTEM = notesSheet['L8'].value
    SOIL_TEXTURE = notesSheet['L9'].value
    INSECTICIDE_RATE = notesSheet['L10'].value
    if INSECTICIDE_RATE != None:
        INSECTICIDE_RATE = INSECTICIDE_RATE
    else:
        INSECTICIDE_RATE = "None"
    FORM_TYPE = "NOTES FORM"

    notesWriteFile.write(GROWER_NAME + "\t" + GROWER_CITY + "\t" + COUNTY + "\t" + ACE_LOCATION + "\t" + STATED_PLOT_ON + "\t" + FLAT_LOCATION + "\t" + str(GPS_LATITUDE) + "\t" + FUNGICIDE + "\t" + CROP + "\t" + str(PLANTING_DATE) + "\t" + str(SEEDING_RATE) + "\t" +
                         str(PLANTING_DEPTH_IN) + "\t" + PLANTER_TYPE + "\t" + str(ROW_WIDTH) + "\t" + str(GPS_LONGITUDE) + "\t" + HERBICIDE + "\t" + PLOT_TYPE + "\t" + IRRIGATION_TYPE + "\t" + PREVIOUS_CROP + "\t" + TILLAGE_SYSTEM + "\t" + SOIL_TEXTURE + "\t" + INSECTICIDE_RATE + "\t" + FORM_TYPE + "\t")


def bottomNotesInfo():
    for row in range(15, notesSheet.max_row + 1):
        HYBRID_VARIETY = notesSheet['A' + str(row)].value
        EMERGENCE = notesSheet['C' + str(row)].value
        if EMERGENCE != None:
            EMERGENCE = EMERGENCE
        else:
            EMERGENCE = "None"
        EARLY_VIGOR = notesSheet['D' + str(row)].value
        if EARLY_VIGOR != None:
            EARLY_VIGOR = EARLY_VIGOR
        else:
            EARLY_VIGOR = "None"
        LATE_PLANT_HEALTH = notesSheet['E' + str(row)].value
        if LATE_PLANT_HEALTH != None:
            LATE_PLANT_HEALTH = LATE_PLANT_HEALTH
        else:
            LATE_PLANT_HEALTH = "None"
        GENERAL_COMMENTS = notesSheet['F' + str(row)].value
        if GENERAL_COMMENTS != None:
            GENERAL_COMMENTS = GENERAL_COMMENTS
        else:
            GENERAL_COMMENTS = "None"
        ENTRY = notesSheet['M' + str(row)].value

        if HYBRID_VARIETY != None and EMERGENCE != None:
            topNotesInfo()
            notesWriteFile.write(str(HYBRID_VARIETY) + "\t" + str(EMERGENCE) + "\t" + str(EARLY_VIGOR) +
                                 "\t" + str(LATE_PLANT_HEALTH) + "\t" + str(GENERAL_COMMENTS) + "\t" + str(GENERAL_COMMENTS) + "\n")
        else:
            continue

    print("Your Notes Form data plot file is done!")


# bottomNotesInfo()
# notesWriteFile.close()


# This is for the HARVEST FORM

# harvestSheet = wb['HARVEST FORM']
# harvestWriteFile = open(
#    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Test Plot(HARVEST).txt', 'w')

def topHarvestInfo():
    GROWER_NAME = harvestSheet['C3'].value
    GROWER_CITY = harvestSheet['C4'].value
    COUNTY = harvestSheet['C5'].value
    ACE_LOCATION = harvestSheet['C6'].value
    STATED_PLOT_ON = harvestSheet['C7'].value
    FLAT_LOCATION = harvestSheet['C8'].value
    GPS_LATITUDE = harvestSheet['C9'].value
    FUNGICIDE = harvestSheet['C10'].value
    if FUNGICIDE != None:
        FUNGICIDE = FUNGICIDE
    else:
        FUNGICIDE = "None"
    HARVEST_DATE = harvestSheet['C11'].value
    if HARVEST_DATE != None:
        HARVEST_DATE = HARVEST_DATE
    else:
        HARVEST_DATE = "None"
    CROP = harvestSheet['H3'].value
    PLANTING_DATE = harvestSheet['H4'].value
    SEEDING_RATE = harvestSheet['H5'].value
    PLANTING_DEPTH_IN = harvestSheet['H6'].value
    PLANTER_TYPE = harvestSheet['H7'].value
    ROW_WIDTH = harvestSheet['H8'].value
    GPS_LONGITUDE = harvestSheet['H9'].value
    HERBICIDE = harvestSheet['H10'].value
    if HERBICIDE != None:
        HERBICIDE = HERBICIDE
    else:
        HERBICIDE = "None"
    COMMODITY_PRICE = harvestSheet['H11'].value
    PLOT_TYPE = harvestSheet['L5'].value
    IRRIGATION_TYPE = harvestSheet['L6'].value
    PREVIOUS_CROP = harvestSheet['L7'].value
    TILLAGE_SYSTEM = harvestSheet['L8'].value
    SOIL_TEXTURE = harvestSheet['L9'].value
    INSECTICIDE_RATE = harvestSheet['L10'].value
    if INSECTICIDE_RATE != None:
        INSECTICIDE_RATE = INSECTICIDE_RATE
    else:
        INSECTICIDE_RATE = "None"
    DRYING_COST = harvestSheet['L11'].value
    FORM_TYPE = "HARVEST FORM"

    harvestWriteFile.write(GROWER_NAME + "\t" + GROWER_CITY + "\t" + COUNTY + "\t" + ACE_LOCATION + "\t" + STATED_PLOT_ON + "\t" + FLAT_LOCATION + "\t" + str(GPS_LATITUDE) + "\t" + FUNGICIDE + "\t" + HARVEST_DATE + "\t" + CROP + "\t" + str(PLANTING_DATE) + "\t" + str(SEEDING_RATE) + "\t" +
                           str(PLANTING_DEPTH_IN) + "\t" + PLANTER_TYPE + "\t" + str(ROW_WIDTH) + "\t" + str(GPS_LONGITUDE) + "\t" + HERBICIDE + "\t" + str(COMMODITY_PRICE) + "\t" + PLOT_TYPE + "\t" + IRRIGATION_TYPE + "\t" + PREVIOUS_CROP + "\t" + TILLAGE_SYSTEM + "\t" + SOIL_TEXTURE + "\t" + INSECTICIDE_RATE + "\t" + str(DRYING_COST) + "\t" + FORM_TYPE + "\t")


def bottomHarvestInfo():
    for row in range(17, harvestSheet.max_row + 1):
        BRAND = harvestSheet['A' + str(row)].value
        PRODUCT = harvestSheet['C' + str(row)].value
        ROW_LENGTH = harvestSheet['F' + str(row)].value
        if ROW_LENGTH != None:
            ROW_LENGTH = ROW_LENGTH
        else:
            ROW_LENGTH = "None"
        WET_WEIGHT = harvestSheet['G' + str(row)].value
        if WET_WEIGHT != None:
            WET_WEIGHT = WET_WEIGHT
        else:
            WET_WEIGHT = "None"
        HARVEST_MOISTURE_PCT = harvestSheet['H' + str(row)].value
        if HARVEST_MOISTURE_PCT != None:
            HARVEST_MOISTURE_PCT = HARVEST_MOISTURE_PCT
        else:
            HARVEST_MOISTURE_PCT = "None"
        NUM_OF_ROWS = harvestSheet['I' + str(row)].value
        TEST_WEIGHT = harvestSheet['J' + str(row)].value
        if TEST_WEIGHT != None:
            TEST_WEIGHT = TEST_WEIGHT
        else:
            TEST_WEIGHT = "None"
        YIELD_PER_ACRE = harvestSheet['K' + str(row)].value
        if YIELD_PER_ACRE != None:
            YIELD_PER_ACRE = YIELD_PER_ACRE
        else:
            YIELD_PER_ACRE = "None"
        PCT_OF_PLOT_ADVANTAGE = harvestSheet['L' + str(row)].value
        if PCT_OF_PLOT_ADVANTAGE != None:
            PCT_OF_PLOT_ADVANTAGE = PCT_OF_PLOT_ADVANTAGE
        else:
            PCT_OF_PLOT_ADVANTAGE = "None"
        YIELD_PER_ACRE_RANK = harvestSheet['M' + str(row)].value
        if YIELD_PER_ACRE_RANK != None:
            YIELD_PER_ACRE_RANK = YIELD_PER_ACRE_RANK
        else:
            YIELD_PER_ACRE_RANK = "None"
        DOLLARS_PER_ACRE_RANK = harvestSheet['N' + str(row)].value
        if DOLLARS_PER_ACRE_RANK != None:
            DOLLARS_PER_ACRE_RANK = DOLLARS_PER_ACRE_RANK
        else:
            DOLLARS_PER_ACRE_RANK = "None"
        if BRAND != None and PRODUCT != None:
            topHarvestInfo()
            harvestWriteFile.write(BRAND + "\t" + str(PRODUCT) + "\t" + str(ROW_LENGTH) +
                                   "\t" + str(WET_WEIGHT) + "\t" + str(HARVEST_MOISTURE_PCT) + "\t" + str(NUM_OF_ROWS) + "\t" + str(TEST_WEIGHT) + "\t" + str(YIELD_PER_ACRE) + "\t" + str(PCT_OF_PLOT_ADVANTAGE) + "\t" + str(YIELD_PER_ACRE_RANK) + "\t" + str(DOLLARS_PER_ACRE_RANK) + "\n")
        else:
            continue

    print("Your Harvest Form data plot file is done!")

# bottomHarvestInfo()
# harvestWriteFile.close()
