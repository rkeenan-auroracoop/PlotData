import os
import pprint
#import openpyxl
from openpyxl import load_workbook

docDir = r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry'

productWriteFile = (
    r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Product Data File(sorted).txt')


def ProductNameReplacement():
    for row in range(17, plantingSheet.max_row + 1):
        COMPANY = plantingSheet['C' + str(row)].value
        HYBRID_VARIETY = plantingSheet['F' + str(row)].value
        if COMPANY != None and HYBRID_VARIETY != None:
            f1.write(str(HYBRID_VARIETY) + "\n")
        else:
            continue


def sorting(filename):
    infile = open(filename)
    words = []
    for line in infile:
        temp = line.split("\n")
        for i in temp:
            words.append(i)
    infile.close()
    words.sort()
    outfile = open(productWriteFile, 'w')
    for i in words:
        outfile.writelines(i)
        outfile.writelines("\n")
    outfile.close()
    print("Your sort is now complete.")


for folders, sub_folders, file in os.walk(docDir):
    for name in file:
        if name.endswith(".xlsx"):
            filename = os.path.join(folders, name)
            print(filename)
            wb = load_workbook(filename)
            plantingSheet = wb['PLANTING FORM']
            with open(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Product Data File.txt', 'a') as f1:
                ProductNameReplacement()
                wb.close()
print("Product Data File is now complete.")

sorting(r'C:\Users\rkeenan\OneDrive - Aurora Cooperative\Documents\2020 Tableau Updates\Al Perry\Product Data File.txt')
